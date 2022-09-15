# -*- coding: utf-8 -*-
import ssl
import re
import urllib.request
import os
from bs4 import BeautifulSoup
import xlwt,http
import openpyxl
import pandas as pd
import sqlite3
# import html5lib
os.mkdir("results-add-1-2-5")

permpath='C:/Users/susu/Desktop/Graduation_Design/FineGrainedAndrodPermissionsStudy/data/安卓权限.xls'
df_sheet = pd.read_excel(permpath,sheet_name = 'manifest.permissions-补充40个',encoding="utf-8",keep_default_na=False)
allpermList =[t for t in df_sheet['Permission name'].tolist() if t]

def main():
    #创建excel,第一列是className，第二列是方法，第三列是tag信息
    # file = open('permissionList.xls',"w")
    # book = xlwt.Workbook() #创建一个Excel
    # sheet1 = book.add_sheet('mapping') #在其中创建一个名为mapping的sheet
    book = openpyxl.Workbook() #创建一个Excel
    sheet1 = book.active #第一个工作表

    # db=

    #循环处理每一个类
    f = open("urlList-add-1-2-5.txt","r")
    content = f.readlines()
    f.close()
    i = 0
    
    
    while i < len(content): #遍历每个url
        url = content[i]
        url = url.replace(".com",".google.cn")
        print(i) 
        print(url)
        '''
        try:
            extractUrl(url,i,sheet1)
        except:
            pass
        '''
        bool = extractUrl(url,i,sheet1) #第i个url,第i个class
        i += 1

    #创建保存文件
    book.save('permissionList-add-1-2-5.xlsx')
    # file.close()
    #extractUrl("https://developer.android.com/reference/android/hardware/camera2/CaptureResult.Key",1)

def findClassName(soup,classCount):
    # print("=================================start findClassName=============================================")

    '''在html里找class名,建立results下的文件夹'''
    metaTag = soup.find_all("meta",attrs={"itemprop": "name"})[0] 
    className = metaTag.get("content")#类名(大标题)
    addclassName = str(className) + "--" +str(classCount)#第classCount个url的类,从0计数
    os.mkdir("results-add-1-2-5" + "/" + addclassName )#每个类建立一个文件夹
    # print("=================================end findClassName=============================================")

    return addclassName


def extractUrl(url,classCount,sheet1):
    # print("=================================start extractUrl=============================================")
    try:
        context = ssl._create_unverified_context()
        content = urllib.request.urlopen(url, context=context).read().decode('utf-8')  #url的html
    except(http.client.IncompleteRead) as e:
        content = e.partial.decode('utf-8')
    except:
        return False
    print("获取页面成功")
    soup = BeautifulSoup(content, 'html5lib')
    className = findClassName(soup,classCount) #在html里找class名,建立results下的文件夹
        
    #写入类名
    #sheet1.write(classCount,0,className.split("--")[0])
    list = soup.find_all("div",attrs={"data-version-added": re.compile("\d")})#每个API的一块描述
    methodCount = 0
    for tag in list:
        try:
            method = tag.h3["id"].split("(")[0]
            # 测试
            # if method=="fields" or method=="Fields":
            # with open('temp.html','a',encoding='utf-8') as f:
            #     f.write(str(tag))
            #     f.write("\n#############################################################\n")
            if  tag.h3["id"][0].isupper() or ("id" in tag.attrs) or ("constants" in method) or ("Constants" in method) or (method=="nested-classes"): #一块里大写的标题，大写是constant
                continue
            else: #小写是method，提取权限
                extractPermission(tag,className,methodCount,sheet1)
        except:
            pass
        methodCount += 1
    return True
    # print("=================================end extractUrl=============================================")

map = {}
count = 1
    

methodSeq = 0 #全局

def extractPermission(tag,className,methodCount,sheet1):
    '''对每个method处理 methodSeq是第几行(全局)
    从method的描述块(tag)里提取权限
    method描述块,该class名，第几个method,写入哪张sheet
    '''

    global methodSeq 
    methodSeq += 1 #下一行

    method = tag.h3["id"].split("(")[0] #实际函数名 如onCurrentVrActivityChanged(android.content.ComponentName)
    print("mothod %d=%s"%(methodSeq,method))
    methodName = str(method) + str(methodCount)#后面加上是第几个函数
    path = "results-add-1-2-5" + "/" + className + "/"+methodName#在类名文件夹下创建函数名文件夹
    os.mkdir(path)
    #print  method

    
    # print("start extractPermission...",end='\t')
    # print("====================================start extractPermission======================================")
    
    temp = ""
    totolPermisionList = []
    children = tag.children#对子节点进行循环

    totolPermisionList = ""
    allPermisionList=""

    for child in children:
        if "permission" in str(child):  #子节点里有permission字符串
            temp = temp+'\n'+ str(child) #子节点整个加入
            logPath ="results-add-1-2-5"+  "/" +className + "/"+ methodName+ "/"+"permission.txt"
            
            f = open(logPath, "a")
            f.write(str(child)) #就把相关子节点整个写入文件夹内"permission.txt做记录
            f.close()
            
            dangerousPermissionSet,allPermissionSet = compareDangerousPermission(child)#返回子节点里有的dangerous权限set

            #去重放到totolPermisionList
            totolPermisionList = addToPermisionList(totolPermisionList,dangerousPermissionSet)#dangerous
            allPermisionList = addToPermisionList(allPermisionList,allPermissionSet)#所有权限
           
    sheet1.cell(methodSeq,1,className)#第一列 类名
    sheet1.cell(methodSeq,2,methodName)#第二列 函数名
    if temp:#有permission的所有子节点
        sheet1.cell(methodSeq,3,str(temp))#第三列 子节点字符串！！！有permission的
    if allPermisionList:
        sheet1.cell(methodSeq,4,str(allPermisionList))#第4列 所有权限
    if totolPermisionList:
        sheet1.cell(methodSeq,5,str(totolPermisionList))#第5列 危险权限(若有)


    map[method] = temp #!!!！！！应该python2.x才可以，是map字典{函数名:[有permission的子节点,]}
    # print("end")
    # print("====================================end extractPermission======================================")


def compareDangerousPermission(child):
    # print("=================================start extractPermission=============================================")

    '''返回子节点里有的dangerous权限set'''
    #所有dangerous，要补一下！！！
    dangerousPermissionList = ["ACCEPT_HANDOVER","ACCESS_COARSE_LOCATION","ACCESS_FINE_LOCATION","ADD_VOICEMAIL","ANSWER_PHONE_CALLS","BODY_SENSORS","CALL_PHONE","CAMERA","GET_ACCOUNTS","PROCESS_OUTGOING_CALLS","READ_CALENDAR","READ_CALL_LOG","READ_CONTACTS","READ_EXTERNAL_STORAGE","READ_PHONE_NUMBERS","READ_PHONE_STATE","READ_SMS","RECEIVE_MMS","RECEIVE_WAP_PUSH","RECORD_AUDIO","SEND_SMS","USE_SIP","WRITE_CALENDAR","WRITE_CALL_LOG","WRITE_CONTACTS","WRITE_EXTERNAL_STORAGE","ACCEPT_HANDOVER","ACTIVITY_RECOGNITION","ACCESS_BACKGROUND_LOCATION","ACCESS_MEDIA_LOCATION","UWB_RANGING","BLUETOOTH_ADVERTISE","BLUETOOTH_CONNECT","BLUETOOTH_SCAN","BODY_SENSORS_BACKGROUND","NEARBY_WIFI_DEVICES","POST_NOTIFICATIONS","READ_MEDIA_AUDIO","READ_MEDIA_IMAGE","READ_MEDIA_VIDEO"]
    # dangerousPermissionList = ["ACCEPT_HANDOVER","ACCESS_COARSE_LOCATION","ACCESS_FINE_LOCATION","ADD_VOICEMAIL","ANSWER_PHONE_CALLS","BODY_SENSORS","CALL_PHONE","CAMERA","GET_ACCOUNTS","PROCESS_OUTGOING_CALLS","READ_CALENDAR","READ_CALL_LOG","READ_CONTACTS","READ_EXTERNAL_STORAGE","READ_PHONE_NUMBERS","READ_PHONE_STATE","READ_SMS","RECEIVE_MMS","RECEIVE_WAP_PUSH","RECORD_AUDIO","SEND_SMS","USE_SIP","WRITE_CALENDAR","WRITE_CALL_LOG","WRITE_CONTACTS","WRITE_EXTERNAL_STORAGE"]
    matchList = set()    
    for dangerousPermission in  dangerousPermissionList:#遍历每个dangerous
        if dangerousPermission in str(child):#如果在子节点里
            matchList.add(dangerousPermission)#加入

    permList = set()
    for perm in  allpermList:#遍历每个dangerous
        if perm in str(child):#如果在子节点里
            permList.add(perm)#加入
    return matchList,permList#返回子节点里有的dangerous
    


def addToPermisionList(totolPermisionList,permissionSet):
    '''危险权限集permissionSet去重 返回list'''
    for p in permissionSet:#遍历这个函数需要的dangerous
        if p not in totolPermisionList:#从permissionSet去重
            totolPermisionList = totolPermisionList+p+";"
    return totolPermisionList #permissionSet去重的list


if __name__ == "__main__":
    main()

        
    

