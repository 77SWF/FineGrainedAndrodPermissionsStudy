import os
#存储数据过大时，会报错
import xlwt 
import openpyxl
import pandas as pd

def addToPermisionList(totolPermisionList,permissionSet):
    '''危险权限集permissionSet去重 返回list'''
    for p in permissionSet:#遍历这个函数需要的dangerous
        if p not in totolPermisionList:#从permissionSet去重
            totolPermisionList.append(p)
    return totolPermisionList #permissionSet去重的list


permpath='C:/Users/susu/Desktop/Graduation_Design/FineGrainedAndrodPermissionsStudy/data/安卓权限.xls'
df_sheet = pd.read_excel(permpath,sheet_name = 'manifest.permissions-补充40个',encoding="utf-8",keep_default_na=False)
allpermList =[t for t in df_sheet['Permission name'].tolist() if t]

#所有dangerous
dangerousPermissionList = ["ACCEPT_HANDOVER","ACCESS_COARSE_LOCATION","ACCESS_FINE_LOCATION","ADD_VOICEMAIL","ANSWER_PHONE_CALLS","BODY_SENSORS","CALL_PHONE","CAMERA","GET_ACCOUNTS","PROCESS_OUTGOING_CALLS","READ_CALENDAR","READ_CALL_LOG","READ_CONTACTS","READ_EXTERNAL_STORAGE","READ_PHONE_NUMBERS","READ_PHONE_STATE","READ_SMS","RECEIVE_MMS","RECEIVE_WAP_PUSH","RECORD_AUDIO","SEND_SMS","USE_SIP","WRITE_CALENDAR","WRITE_CALL_LOG","WRITE_CONTACTS","WRITE_EXTERNAL_STORAGE","ACCEPT_HANDOVER","ACTIVITY_RECOGNITION","ACCESS_BACKGROUND_LOCATION","ACCESS_MEDIA_LOCATION","UWB_RANGING","BLUETOOTH_ADVERTISE","BLUETOOTH_CONNECT","BLUETOOTH_SCAN","BODY_SENSORS_BACKGROUND","NEARBY_WIFI_DEVICES","POST_NOTIFICATIONS","READ_MEDIA_AUDIO","READ_MEDIA_IMAGE","READ_MEDIA_VIDEO"]



#创建excel,第一列是className，第二列是方法，第三列是tag信息
# file = open('permissionList400.xlsx',"w")
book = openpyxl.Workbook() #创建一个Excel
# sheet1 = book.add_sheet('mapping') 
sheet1 = book.active #第一个工作表

path = './results-25'
class_list = os.listdir(path)
#按时间排序
class_list = sorted(class_list,key=lambda x: os.path.getmtime(os.path.join(path, x)))
# print(class_list)

methodSeq = 0
num_class = 0
for class_name in class_list:
    num_class += 1
    print("\rclass: %d/577"%num_class,end="")
    class_path = os.path.join(path,class_name)
    method_list = os.listdir(class_path)
    if(method_list): #类内有函数
        method_list = sorted(method_list,key=lambda x: os.path.getmtime(os.path.join(class_path, x)))
        
        #遍历每个method
        for method_name in method_list:
            methodSeq +=1
            # print("第%d个method"%(methodSeq+1))
            method_path = os.path.join(class_path,method_name)
            perm_txt_list = os.listdir(method_path)
            if(perm_txt_list):#有permission.txt
                perm_txt_path = os.path.join(method_path,perm_txt_list[0])
                with open(perm_txt_path,'r',encoding='utf-8') as f:
                    child = str(f.read())
                    # print(child)
            else:
                child = ""

            dangerousPermissionSet = set()    
            for dangerousPermission in  dangerousPermissionList:#遍历每个dangerous
                if dangerousPermission in str(child):#如果在子节点里
                    dangerousPermissionSet.add(dangerousPermission)#加入

            allPermissionSet = set()
            for perm in  allpermList:#遍历每个dangerous
                if perm in str(child):#如果在子节点里
                    allPermissionSet.add(perm)#加入

            totolPermisionList=[]
            allPermisionList=[]
            totolPermisionList = addToPermisionList(totolPermisionList,dangerousPermissionSet)#dangerous
            allPermisionList = addToPermisionList(allPermisionList,allPermissionSet)#所有权限

            sheet1.cell(methodSeq,1,class_name)#第一列 类名
            sheet1.cell(methodSeq,2,method_name)#第二列 函数名
            if child:#有permission的子节点
                sheet1.cell(methodSeq,3,str(child))#第三列 子节点字符串！！！有permission的
            if allPermisionList:
                sheet1.cell(methodSeq,4,str(allPermisionList))#第4列 所有权限
            if totolPermisionList:
                sheet1.cell(methodSeq,5,str(totolPermisionList))#第5列 危险权限(若有)

    
    else:#类内无函数
        continue
        
            
book.save('permissionList.xlsx')
# file.close()